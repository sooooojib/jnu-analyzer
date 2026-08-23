"""
Student Academic Analysis Excel Exporter using openpyxl.

Builds a professional, multi-tab Excel workbook (.xlsx) strictly scoped
to the selected student using data from build_student_report_data.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def build_student_excel(report_data: Dict[str, Any]) -> bytes:
    """
    Generates a structured, multi-sheet Excel workbook for student analysis.

    Args:
        report_data: Dict output from build_student_report_data()

    Returns:
        Raw XLSX file bytes.
    """
    wb = Workbook()

    # Style definitions
    font_title = Font(name="Calibri", size=14, bold=True, color="0F172A")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="475569")
    font_section = Font(name="Calibri", size=11, bold=True, color="0F172A")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
    font_normal = Font(name="Calibri", size=10, color="1E293B")
    font_meta_label = Font(name="Calibri", size=10, bold=True, color="334155")
    font_emerald = Font(name="Calibri", size=10, bold=True, color="059669")

    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_section = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_kpi_card = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    metadata = report_data.get("metadata", {})
    student_info = report_data.get("student_info", {})
    academic = report_data.get("academic_summary", {})
    subjects = report_data.get("subject_results", [])
    stats = report_data.get("performance_statistics", {})

    # =========================================================================
    # SHEET 1: Student Summary
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Student Summary"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner
    ws1["A1"] = "JnU Analyzer — Student Academic Analysis"
    ws1["A1"].font = font_title
    inst_info = f"{metadata.get('institution', 'Jagannath University')} • {metadata.get('department', 'Department of Computer Science & Engineering')}"
    ws1["A2"] = inst_info
    ws1["A2"].font = font_subtitle
    if metadata.get("semester"):
        sem_str = f"Semester: {metadata.get('semester')}"
        if metadata.get("exam_session"):
            sem_str += f" | Session: {metadata.get('exam_session')}"
        ws1["A3"] = sem_str
        ws1["A3"].font = font_subtitle

    # Section 1: Student Identification
    ws1["A5"] = "Student Identification"
    ws1["A5"].font = font_section
    ws1["A5"].fill = fill_section
    ws1.merge_cells("A5:D5")

    id_rows = [
        ("Student Name", str(student_info.get("student_name", "UNKNOWN")), "Student ID", str(student_info.get("student_id", "N/A"))),
        ("Report Date", datetime.now().strftime("%Y-%m-%d %H:%M"), "", ""),
    ]

    for row_idx, (l1, v1, l2, v2) in enumerate(id_rows, start=6):
        ws1[f"A{row_idx}"] = l1
        ws1[f"A{row_idx}"].font = font_meta_label
        ws1[f"B{row_idx}"] = v1
        ws1[f"B{row_idx}"].font = font_bold

        ws1[f"C{row_idx}"] = l2
        ws1[f"C{row_idx}"].font = font_meta_label
        ws1[f"D{row_idx}"] = v2
        ws1[f"D{row_idx}"].font = font_normal

        for col_letter in ["A", "B", "C", "D"]:
            ws1[f"{col_letter}{row_idx}"].border = thin_border

    # Section 2: Academic Performance Summary
    ws1["A9"] = "Academic Performance Summary"
    ws1["A9"].font = font_section
    ws1["A9"].fill = fill_section
    ws1.merge_cells("A9:D9")

    sem_rank_text = f"#{academic.get('semester_rank')}" if academic.get("semester_rank") else "—"
    if academic.get("semester_percentile"):
        sem_rank_text += f" (Top {100 - academic.get('semester_percentile'):.1f}%)"

    cum_rank_text = f"#{academic.get('cumulative_rank')}" if academic.get("cumulative_rank") else "—"
    if academic.get("cumulative_percentile"):
        cum_rank_text += f" (Top {100 - academic.get('cumulative_percentile'):.1f}%)"

    summary_rows = [
        ("Current Semester GPA", academic.get("semester_gpa", 0.0), "Semester Result Status", academic.get("semester_result_status", "PASSED")),
        ("Current Semester Rank", sem_rank_text, "Credits Attempted", academic.get("credits_attempted", 0.0)),
        ("Cumulative CGPA", academic.get("cumulative_cgpa", 0.0) if academic.get("cumulative_cgpa", 0.0) > 0 else "N/A", "Cumulative Result Status", academic.get("cumulative_result_status", "PASSED")),
        ("Cumulative Rank", cum_rank_text, "Credits Earned", academic.get("credits_earned", 0.0)),
        ("Total Subjects Analyzed", academic.get("total_subjects", len(subjects)), "Remarks", academic.get("semester_remarks") or "—"),
    ]

    for row_idx, (l1, v1, l2, v2) in enumerate(summary_rows, start=10):
        ws1[f"A{row_idx}"] = l1
        ws1[f"A{row_idx}"].font = font_meta_label
        ws1[f"B{row_idx}"] = v1
        ws1[f"B{row_idx}"].font = font_emerald if "GPA" in l1 or "CGPA" in l1 else font_bold
        if isinstance(v1, float):
            ws1[f"B{row_idx}"].number_format = "0.00"

        ws1[f"C{row_idx}"] = l2
        ws1[f"C{row_idx}"].font = font_meta_label
        ws1[f"D{row_idx}"] = v2
        ws1[f"D{row_idx}"].font = font_normal
        if isinstance(v2, float):
            ws1[f"D{row_idx}"].number_format = "0.0"

        for col_letter in ["A", "B", "C", "D"]:
            ws1[f"{col_letter}{row_idx}"].border = thin_border

    # Auto-adjust column widths for Sheet 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # =========================================================================
    # SHEET 2: Subject Results
    # =========================================================================
    ws2 = wb.create_sheet(title="Subject Results")
    ws2.views.sheetView[0].showGridLines = True

    headers_s2 = [
        "Course Code",
        "Course Title",
        "Credit",
        "Grade Point",
        "Letter Grade",
        "Subject Rank",
        "Class Average GP",
        "Status",
    ]

    for col_idx, h in enumerate(headers_s2, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for row_idx, sub in enumerate(subjects, start=2):
        gp_val = sub.get("grade_point")
        avg_gp = sub.get("class_average_gp")
        is_even = (row_idx % 2 == 0)
        row_fill = fill_zebra if is_even else None

        ws2.cell(row=row_idx, column=1, value=sub.get("course_code", "")).alignment = align_center
        ws2.cell(row=row_idx, column=2, value=sub.get("course_title", "")).alignment = align_wrap
        
        c_cell = ws2.cell(row=row_idx, column=3, value=sub.get("credit", 3.0))
        c_cell.alignment = align_center
        c_cell.number_format = "0.0"

        gp_cell = ws2.cell(row=row_idx, column=4, value=gp_val if gp_val is not None else "")
        gp_cell.alignment = align_center
        gp_cell.font = font_emerald
        if gp_val is not None:
            gp_cell.number_format = "0.00"

        ws2.cell(row=row_idx, column=5, value=sub.get("letter_grade", "")).alignment = align_center
        ws2.cell(row=row_idx, column=6, value=f"#{sub.get('subject_rank')}" if sub.get("subject_rank") else "—").alignment = align_center
        
        avg_cell = ws2.cell(row=row_idx, column=7, value=avg_gp if avg_gp is not None else "")
        avg_cell.alignment = align_center
        if avg_gp is not None:
            avg_cell.number_format = "0.00"

        ws2.cell(row=row_idx, column=8, value=sub.get("status", "VALID")).alignment = align_center

        for col_idx in range(1, 9):
            c = ws2.cell(row=row_idx, column=col_idx)
            c.border = thin_border
            c.font = font_normal if col_idx != 4 else font_emerald
            if row_fill:
                c.fill = row_fill

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:H{len(subjects) + 1}"

    # Auto-adjust column widths for Sheet 2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        if col_letter == "B":
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 35)
        else:
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # =========================================================================
    # SHEET 3: Student Statistics
    # =========================================================================
    ws3 = wb.create_sheet(title="Student Statistics")
    ws3.views.sheetView[0].showGridLines = True

    ws3["A1"] = "Student Performance Analytics"
    ws3["A1"].font = font_title
    ws3["A2"] = f"Deterministic insights computed for {student_info.get('student_name', 'Student')} ({student_info.get('student_id', 'N/A')})"
    ws3["A2"].font = font_subtitle

    ws3["A4"] = "Performance Insights"
    ws3["A4"].font = font_section
    ws3["A4"].fill = fill_section
    ws3.merge_cells("A4:C4")

    highest_courses = ", ".join(stats.get("highest_subject_courses", [])) or "—"
    lowest_courses = ", ".join(stats.get("lowest_subject_courses", [])) or "—"

    stat_rows = [
        ("Highest Subject GP", stats.get("highest_subject_gp"), f"Achieved in {highest_courses}"),
        ("Lowest Subject GP", stats.get("lowest_subject_gp"), f"In {lowest_courses}"),
        ("Average Subject GP", stats.get("average_subject_gp"), "Arithmetic mean of all course grade points"),
        ("Semester GPA", academic.get("semester_gpa", 0.0), f"Credit-weighted semester average ({academic.get('semester_result_status', 'PASSED')})"),
        ("Cumulative CGPA", academic.get("cumulative_cgpa", 0.0), "Extracted cumulative grade point average"),
        ("Class Standing", f"Rank #{academic.get('semester_rank', '—')}", f"Percentile: {academic.get('semester_percentile', '—')}%"),
    ]

    for row_idx, (label, val, desc) in enumerate(stat_rows, start=5):
        ws3[f"A{row_idx}"] = label
        ws3[f"A{row_idx}"].font = font_meta_label

        ws3[f"B{row_idx}"] = val if val is not None else "—"
        ws3[f"B{row_idx}"].font = font_emerald if isinstance(val, float) else font_bold
        ws3[f"B{row_idx}"].alignment = align_center
        if isinstance(val, float):
            ws3[f"B{row_idx}"].number_format = "0.00"

        ws3[f"C{row_idx}"] = desc
        ws3[f"C{row_idx}"].font = font_normal

        for col_letter in ["A", "B", "C"]:
            ws3[f"{col_letter}{row_idx}"].border = thin_border

    # Auto-adjust column widths for Sheet 3
    for col in ws3.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(max_len + 4, 24)

    # Save to in-memory bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
